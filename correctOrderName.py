import openpyxl

names = []
newNames = []

def ReadFile(): # 1 -- read the excel file
    theFile = openpyxl.load_workbook( 'users.xlsx' ) # the excel file is the same path of the python file
    allSheetNames = theFile.sheetnames

    print( "===========================================" )
    print( "Extracting data from {}".format( theFile.sheetnames ) )
    print( "===========================================" )

    for sheet in allSheetNames:
        currentSheet = theFile[sheet]
        for row in range( 1, currentSheet.max_row + 1 ):
            for column in "A":  #  if I add more letter, I add more columns
                cell_name = "{}{}".format( column, row )           
                LoadNames( currentSheet[cell_name].value )

def LoadNames( name ): # 2 -- load the names in the new array
    if name != "":
        names.append( name )

def ChangeOrderRow(): # 3 --- change the names order
    for n in names:
        if n != None:
            ultraSplit = " ".join(n.split()) # delete double spaces
            splitName = ultraSplit.split( ' ' )
            if len(splitName) == 4:
                LoadNewNames( splitName[2].capitalize() + ' ' + splitName[3].capitalize()  + ' ' + splitName[0].capitalize()  + ' ' + splitName[1].capitalize() )
            elif len(splitName) == 3:
                LoadNewNames( splitName[2].capitalize() + ' ' + splitName[0].capitalize()  + ' ' + splitName[1].capitalize() )
            elif len(splitName) == 2:
                LoadNewNames( splitName[1].capitalize() + ' ' + splitName[0].capitalize() )
            else:
                print( "Some users or one user is incorrect format" )

def LoadNewNames( name ): # 4 -- load the names with correct format
    if name != "":
        newNames.append( name )





def PrintNames(): # ++00 -- print
    for n in newNames:
        print( n )
    print( "===========================================" )

# ------------------------------------------------- execution -------------------------------------------------
ReadFile()
ChangeOrderRow()
PrintNames()
